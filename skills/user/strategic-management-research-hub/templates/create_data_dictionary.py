#!/usr/bin/env python3
"""
データ辞書テンプレート作成スクリプト
Data Dictionary Template Creator
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ワークブック作成
wb = Workbook()

# シート1: 変数一覧 (Variable List)
ws1 = wb.active
ws1.title = "Variables"

# ヘッダー行のスタイル設定
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ヘッダー
headers = [
    "Variable ID",
    "Variable Name (English)",
    "Variable Name (Japanese)",
    "Variable Type",
    "Definition",
    "Measurement",
    "Unit",
    "Data Type",
    "Source",
    "Source URL",
    "Coverage Period",
    "Missing Data %",
    "Transformation",
    "Notes"
]

ws1.append(headers)

# ヘッダー行のスタイル適用
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# サンプルデータ行
sample_data = [
    [
        "DV01",
        "ROA",
        "総資産利益率",
        "Dependent Variable",
        "Return on Assets = Net Income / Total Assets",
        "Net Income divided by Total Assets",
        "%",
        "Continuous",
        "Compustat / EDINET",
        "https://wrds-www.wharton.upenn.edu/",
        "1980-2024",
        "2.3%",
        "Winsorized at 1%/99%",
        "Primary performance measure"
    ],
    [
        "IV01",
        "R&D Intensity",
        "R&D強度",
        "Independent Variable",
        "R&D Intensity = R&D Expenditure / Sales Revenue",
        "R&D spending as % of sales",
        "%",
        "Continuous",
        "Compustat",
        "https://wrds-www.wharton.upenn.edu/",
        "1980-2024",
        "42.7%",
        "Missing = 0 (non-R&D firms)",
        "R&D data not reported by all firms"
    ],
    [
        "MOD01",
        "Environmental Dynamism",
        "環境動態性",
        "Moderator",
        "Standard deviation of industry sales growth over 5-year rolling window",
        "σ(Industry Sales Growth, t-4 to t)",
        "Decimal",
        "Continuous",
        "Compustat",
        "https://wrds-www.wharton.upenn.edu/",
        "1985-2024",
        "0.5%",
        "Calculated from industry data",
        "Measures environmental uncertainty"
    ],
    [
        "CV01",
        "Firm Size",
        "企業規模",
        "Control Variable",
        "Natural logarithm of total assets",
        "log(Total Assets)",
        "log($)",
        "Continuous",
        "Compustat / EDINET",
        "https://wrds-www.wharton.upenn.edu/",
        "1980-2024",
        "0.1%",
        "Log transformation",
        "Controls for scale effects"
    ],
    [
        "CV02",
        "Firm Age",
        "企業年齢",
        "Control Variable",
        "Number of years since firm founding",
        "Current Year - Founding Year",
        "Years",
        "Count",
        "Compustat / EDINET",
        "https://wrds-www.wharton.upenn.edu/",
        "1980-2024",
        "1.2%",
        "None",
        "Founding year from incorporation date"
    ],
    [
        "CV03",
        "Leverage",
        "レバレッジ",
        "Control Variable",
        "Total Debt / Total Assets",
        "Sum of short-term and long-term debt divided by total assets",
        "Ratio (0-1)",
        "Continuous",
        "Compustat / EDINET",
        "https://wrds-www.wharton.upenn.edu/",
        "1980-2024",
        "0.3%",
        "Winsorized at 1%/99%",
        "Financial leverage measure"
    ]
]

for row_data in sample_data:
    ws1.append(row_data)
    for cell in ws1[ws1.max_row]:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# 列幅の調整
column_widths = {
    'A': 12, 'B': 20, 'C': 20, 'D': 18, 'E': 35, 'F': 30,
    'G': 10, 'H': 12, 'I': 20, 'J': 40, 'K': 15, 'L': 12, 'M': 25, 'N': 40
}
for col, width in column_widths.items():
    ws1.column_dimensions[col].width = width

# 行の高さ
ws1.row_dimensions[1].height = 40

# シート2: Data Sources
ws2 = wb.create_sheet("Data Sources")

headers2 = [
    "Source ID",
    "Source Name",
    "Provider",
    "Access Method",
    "Cost",
    "Coverage",
    "Geographic Scope",
    "Temporal Scope",
    "Update Frequency",
    "URL",
    "Contact",
    "Terms of Use",
    "Citation"
]

ws2.append(headers2)

for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

sample_sources = [
    [
        "DS01",
        "Compustat North America",
        "S&P Global Market Intelligence",
        "WRDS Subscription",
        "Paid (University License)",
        "10,000+ US/Canadian firms",
        "North America",
        "1950-Present",
        "Quarterly",
        "https://wrds-www.wharton.upenn.edu/",
        "wrds@wharton.upenn.edu",
        "Academic use only",
        "Compustat via WRDS (accessed [date])"
    ],
    [
        "DS02",
        "EDINET",
        "Financial Services Agency Japan",
        "Public Website Download",
        "Free",
        "All Japanese listed firms",
        "Japan",
        "2008-Present",
        "Annual/Quarterly",
        "https://disclosure2.edinet-fsa.go.jp/",
        "N/A",
        "Public domain",
        "EDINET (Financial Services Agency Japan, accessed [date])"
    ],
    [
        "DS03",
        "USPTO PatentsView",
        "U.S. Patent and Trademark Office",
        "API / Bulk Download",
        "Free",
        "All US patents",
        "United States (global applicants)",
        "1976-Present",
        "Weekly",
        "https://patentsview.org/",
        "patentsview@air.org",
        "Public domain",
        "USPTO PatentsView (accessed [date])"
    ]
]

for row_data in sample_sources:
    ws2.append(row_data)
    for cell in ws2[ws2.max_row]:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

column_widths2 = {
    'A': 10, 'B': 25, 'C': 30, 'D': 20, 'E': 20, 'F': 25,
    'G': 18, 'H': 15, 'I': 15, 'J': 40, 'K': 25, 'L': 20, 'M': 40
}
for col, width in column_widths2.items():
    ws2.column_dimensions[col].width = width

ws2.row_dimensions[1].height = 40

# シート3: Variable Types
ws3 = wb.create_sheet("Variable Types")

headers3 = [
    "Variable Type",
    "Abbreviation",
    "Description",
    "Example"
]

ws3.append(headers3)

for cell in ws3[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

type_data = [
    ["Dependent Variable", "DV", "Outcome variable that the study aims to explain", "ROA, Innovation Output, Market Share"],
    ["Independent Variable", "IV", "Predictor variable hypothesized to cause changes in DV", "R&D Intensity, Board Diversity, Alliance Portfolio Size"],
    ["Moderator", "MOD", "Variable that affects the strength/direction of IV-DV relationship", "Environmental Dynamism, Firm Size"],
    ["Mediator", "MED", "Variable that explains the mechanism between IV and DV", "Absorptive Capacity, Organizational Learning"],
    ["Control Variable", "CV", "Variable included to rule out alternative explanations", "Firm Size, Firm Age, Industry, Year"],
    ["Instrumental Variable", "INS", "Exogenous variable used to address endogeneity", "Distance to Patent Office, Regulatory Change"]
]

for row_data in type_data:
    ws3.append(row_data)
    for cell in ws3[ws3.max_row]:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

column_widths3 = {'A': 20, 'B': 15, 'C': 50, 'D': 50}
for col, width in column_widths3.items():
    ws3.column_dimensions[col].width = width

ws3.row_dimensions[1].height = 40

# シート4: Data Types
ws4 = wb.create_sheet("Data Types")

headers4 = [
    "Data Type",
    "Description",
    "Example Values",
    "Statistical Method"
]

ws4.append(headers4)

for cell in ws4[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

data_type_info = [
    ["Continuous", "Numeric variable that can take any value within a range", "ROA = 5.34%, R&D Intensity = 3.2%", "OLS, Panel Regression"],
    ["Count", "Non-negative integer representing counts", "Patent Count = 15, Alliance Count = 3", "Poisson, Negative Binomial"],
    ["Binary", "Categorical variable with two outcomes", "Innovation = 1 (yes) or 0 (no)", "Logistic Regression, Probit"],
    ["Ordinal", "Categorical variable with ordered levels", "Rating = 1 (low) to 5 (high)", "Ordered Logit/Probit"],
    ["Nominal", "Categorical variable with unordered categories", "Industry = Manufacturing, Services, Tech", "Dummy Variables, Multinomial Logit"],
    ["Ratio", "Continuous variable with meaningful zero", "Sales = $100M, Assets = $500M", "OLS, Log Transformation"],
    ["Percentage", "Continuous variable expressed as proportion", "Market Share = 15%, Female Directors = 30%", "Logit Transformation, Beta Regression"]
]

for row_data in data_type_info:
    ws4.append(row_data)
    for cell in ws4[ws4.max_row]:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

column_widths4 = {'A': 15, 'B': 40, 'C': 40, 'D': 30}
for col, width in column_widths4.items():
    ws4.column_dimensions[col].width = width

ws4.row_dimensions[1].height = 40

# シート5: Transformations
ws5 = wb.create_sheet("Transformations")

headers5 = [
    "Transformation",
    "Purpose",
    "Formula",
    "When to Use",
    "Example"
]

ws5.append(headers5)

for cell in ws5[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

transformation_data = [
    [
        "Log Transformation",
        "Handle skewness, reduce influence of outliers",
        "Y' = log(Y)",
        "When variable is highly skewed (e.g., firm size, sales)",
        "log(Total Assets)"
    ],
    [
        "Winsorization",
        "Cap extreme values to reduce outlier impact",
        "Replace values <1% with 1st percentile, >99% with 99th percentile",
        "For continuous DVs (ROA, ROE) to handle outliers",
        "ROA: cap at 1%/99%"
    ],
    [
        "Standardization",
        "Mean-center and scale for interaction terms",
        "Z = (X - μ) / σ",
        "Before creating interaction terms to reduce multicollinearity",
        "(R&D - mean) / SD"
    ],
    [
        "Lagging",
        "Address temporal ordering, reduce endogeneity",
        "Y(t) ~ X(t-1)",
        "When IV may be endogenous to DV",
        "ROA(t) ~ R&D(t-1)"
    ],
    [
        "Difference",
        "Control for time-invariant unobserved heterogeneity",
        "ΔY = Y(t) - Y(t-1)",
        "In first-difference models or DiD designs",
        "ΔSales = Sales(t) - Sales(t-1)"
    ],
    [
        "Indicator Method",
        "Handle missing data",
        "Create dummy for missing + impute with 0 or mean",
        "When data is not missing at random (e.g., R&D)",
        "R&D_missing = 1/0, R&D = 0 if missing"
    ]
]

for row_data in transformation_data:
    ws5.append(row_data)
    for cell in ws5[ws5.max_row]:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

column_widths5 = {'A': 18, 'B': 35, 'C': 35, 'D': 40, 'E': 25}
for col, width in column_widths5.items():
    ws5.column_dimensions[col].width = width

ws5.row_dimensions[1].height = 40

# シート6: Instructions
ws6 = wb.create_sheet("Instructions")

instructions = [
    ["Data Dictionary Template - 使用方法 / Instructions", ""],
    ["", ""],
    ["目的 / Purpose:", "このテンプレートは、戦略・組織論研究における変数を体系的に記録し、データの再現性を確保するためのものです。"],
    ["", "This template systematically documents variables in strategy and organization research to ensure data reproducibility."],
    ["", ""],
    ["使い方 / How to Use:", ""],
    ["1. 「Variables」シート", "すべての変数（従属変数、独立変数、調整変数、統制変数）を記録します。サンプルデータを参考に、各列を埋めてください。"],
    ["   'Variables' Sheet", "Record all variables (DV, IV, MOD, CV). Refer to sample data and fill in each column."],
    ["", ""],
    ["2. 「Data Sources」シート", "使用するデータソースの詳細を記録します。引用に必要な情報を含めてください。"],
    ["   'Data Sources' Sheet", "Document details of data sources used. Include information needed for citation."],
    ["", ""],
    ["3. 「Variable Types」シート", "変数の種類（DV, IV, MOD, MED, CV）の定義を確認できます。（編集不要）"],
    ["   'Variable Types' Sheet", "Reference sheet for variable type definitions. (No editing needed)"],
    ["", ""],
    ["4. 「Data Types」シート", "データ型（連続、カウント、バイナリ等）の説明を確認できます。（編集不要）"],
    ["   'Data Types' Sheet", "Reference sheet for data type descriptions. (No editing needed)"],
    ["", ""],
    ["5. 「Transformations」シート", "データ変換方法（対数変換、ウィンソライズ等）の参考情報です。（編集不要）"],
    ["   'Transformations' Sheet", "Reference sheet for data transformation methods. (No editing needed)"],
    ["", ""],
    ["ベストプラクティス / Best Practices:", ""],
    ["✓", "すべての変数に一意のIDを付与する（DV01, IV01, CV01等）"],
    ["✓", "変数の定義と測定方法を明確に記述する"],
    ["✓", "データソースのURLを記録する（再現性のため）"],
    ["✓", "欠損値の割合を記録する"],
    ["✓", "データ変換方法を明記する（対数変換、ウィンソライズ等）"],
    ["✓", "研究ノートに重要な注意事項を記載する"],
    ["", ""],
    ["再現性のチェックリスト / Reproducibility Checklist:", ""],
    ["□", "すべての変数の定義が明確か？"],
    ["□", "データソースが特定可能か？"],
    ["□", "変数構築の手順が再現可能か？"],
    ["□", "欠損値の処理方法が記録されているか？"],
    ["□", "外れ値の処理方法が記録されているか？"],
    ["", ""],
    ["Citation Example:", ""],
    ["", "「変数の操作化とデータソースの詳細はData Dictionary（Appendix A）を参照」"],
    ["", "\"See Data Dictionary (Appendix A) for variable operationalization and data sources.\""],
]

for row_data in instructions:
    ws6.append(row_data)

# 1列目をボールド
for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(bold=True)

# タイトル行（1行目）を大きく、青背景
ws6['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws6['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws6.merge_cells('A1:B1')
ws6['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 30

# 列幅
ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 100

# すべてのセルにワードラップ
for row in ws6.iter_rows(min_row=1, max_row=ws6.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# ファイル保存
output_path = "/home/claude/strategic-research-enhancement/templates/data_dictionary.xlsx"
wb.save(output_path)

print(f"Data Dictionary Template created: {output_path}")
print(f"Sheets created: {', '.join(wb.sheetnames)}")
print("✓ Variables sheet with sample data")
print("✓ Data Sources sheet")
print("✓ Variable Types reference")
print("✓ Data Types reference")
print("✓ Transformations reference")
print("✓ Instructions sheet (Japanese/English)")
